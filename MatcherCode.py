import streamlit as st
import pandas as pd
import difflib
from openpyxl import load_workbook
import io

st.set_page_config(page_title="Excel Auto-Filler", page_icon="📘", layout="wide")
st.title("📘 Excel Auto-Filler from Master Dataset")
st.markdown("""
Upload:
1. **Master Excel file** (must contain *Assays* and *QA* sheets)  
2. **Unfilled Excel file** (to auto-fill data under question columns)
""")

# ------------------ UPLOAD FILES ------------------
master_file = st.file_uploader("Upload Master File (with 'Assays' & 'QA' sheets)", type=["xlsx"])
unfilled_file = st.file_uploader("Upload Excel to Fill", type=["xlsx"])

if master_file and unfilled_file:
    with st.spinner("Processing..."):
        # ------------------ PARAMETERS ------------------
        ASSAYS_SHEET = "Assays"
        QA_SHEET = "QA"
        QUESTION_ROW = 6
        DATA_START_ROW = QUESTION_ROW + 1
        PARAMETER_COL = 2
        START_COL = 7

        # ------------------ LOAD MASTER ------------------
        master_df = pd.read_excel(master_file, sheet_name=ASSAYS_SHEET, engine='openpyxl')
        master_df.columns = master_df.columns.str.strip().str.lower().str.replace(" ", "_")

        required_cols = ["product_name", "parameter", "test_name_id", "alias"]
        for c in required_cols:
            if c not in master_df.columns:
                master_df[c] = ""

        # Alias map and test info
        alias_map = {}
        for _, row in master_df.iterrows():
            names = [
                str(row["test_name_id"]).strip(),
                str(row["product_name"]).strip(),
                str(row["parameter"]).strip(),
                str(row["alias"]).strip(),
            ]
            for alias in names:
                if alias and alias.lower() not in alias_map:
                    alias_map[alias.lower()] = str(row["test_name_id"]).strip()

        test_info = {str(row["test_name_id"]).strip(): row.to_dict() for _, row in master_df.iterrows()}

        # ------------------ LOAD QA SHEET ------------------
        qa_df = pd.read_excel(master_file, sheet_name=QA_SHEET, header=None)
        qa_questions = qa_df.iloc[0].dropna().astype(str).tolist()
        qa_answers = qa_df.iloc[1].dropna().astype(str).tolist()

        # ------------------ FUZZY MATCH ------------------
        def find_best_test_match(text, cutoff=0.6):
            text = str(text).strip().lower()
            if not text:
                return None
            match = difflib.get_close_matches(text, list(alias_map.keys()), n=1, cutoff=cutoff)
            return alias_map[match[0]] if match else None

        # ------------------ LOAD UNFILLED ------------------
        wb = load_workbook(unfilled_file)
        ws = wb.active
        filled_count = 0
        unmatched_params = []
        unmatched_questions = []

        # ------------------ MAIN LOGIC ------------------
        for col in range(START_COL, ws.max_column + 1):
            question = ws.cell(row=QUESTION_ROW, column=col).value
            if not question or not str(question).strip():
                continue

            match_q = difflib.get_close_matches(str(question).strip().lower(),
                                                [q.lower().strip() for q in qa_questions], n=1, cutoff=0.6)
            if not match_q:
                unmatched_questions.append(question)
                continue

            q_idx = [q.lower().strip() for q in qa_questions].index(match_q[0])
            print(q_inx)
            answer_col = qa_answers[q_idx].strip().lower()

            if answer_col not in master_df.columns:
                unmatched_questions.append(f"{question} → {answer_col} (not in master)")
                continue

            for row in range(DATA_START_ROW, ws.max_row + 1):
                param = ws.cell(row=row, column=PARAMETER_COL).value
                if not param:
                    break
                cell = ws.cell(row=row, column=col)
                if cell.value and str(cell.value).strip():
                    continue
                matched_test = find_best_test_match(param)
                if not matched_test:
                    unmatched_params.append(param)
                    continue
                test_row = test_info.get(matched_test)
                if not test_row:
                    unmatched_params.append(param)
                    continue
                value = test_row.get(answer_col, "")
                if pd.notna(value):
                    cell.value = str(value)
                    filled_count += 1

        # ------------------ SAVE TO BUFFER ------------------
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        base_name = unfilled_file.name.rsplit(".", 1)[0]
        output_name = f"{base_name}_FILLED.xlsx"

        # ------------------ SHOW RESULTS ------------------
        st.success(f"✅ Auto-fill complete. {filled_count} cells filled.")
        st.download_button(
            label="📥 Download Filled Excel",
            data=output,
            file_name=output_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        if unmatched_questions:
            st.warning(f"⚠️ Unmatched Questions ({len(unmatched_questions)}):")
            st.write(unmatched_questions[:15])

        if unmatched_params:
            st.warning(f"⚠️ Unmatched Parameters:")
            st.write(unmatched_params[:15])

else:
    st.info("⬆️ Please upload both the Master file and the unfilled Excel file to start.")
